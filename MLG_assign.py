#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @package MLGassign.py
# @author Sebastien Ravel

"""
    :author: Sebastien Ravel
    :contact: sebastien.ravel@cirad.fr
    :date: 25/02/2019
    :version: 1.0

    Script description
    ------------------
"""
##################################################
# Modules
##################################################
# Python modules
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from gooey import Gooey, GooeyParser
# remove pandas header
# styles
# this avoids the restriction that xlsxwriter cannot
# format cells where formatting was already applied
import pandas.io.formats.excel

# pandas.io.formats.excel.header_style = None
pandas.set_option('display.width', 1000)
pandas.set_option('display.precision', 0)  # formate le nombre de chiffre après la virgule pour tout les dataframes

##################################################
# Variables Globales

__version__ = "1.0.3"
epilogTools = ""

descriptionTools = f"""
{'#' * 80}
# Add MLG on excel file with table like :
{'_' * 67}
|{'':^10}|{'Pymrs47':^10}|{'Pyrms427':^10}|{'Pyrms657':^10}|{'Pyrms77':^10}|{'Pyrms63':^10}|
{'-' * 67}
|{'MD2290':^10}|{'163':^10}|{'211':^10}|{'168':^10}|{'194':^10}|{'151':^10}|
|{'FR6334':^10}|{'164':^10}|{'211':^10}|{'168':^10}|{'194':^10}|{'155':^10}|
{'_' * 67}

#\tIntellectual property belongs to CIRAD BGPI - Written by Sébastien Ravel
#\t{__version__}
{'#' * 80}
"""

LICENSE = open("LICENSE","r").read()

##################################################
# Functions

def sort_human(s, _nsre=re.compile('([0-9]+)')):
    """ Sort the list in the way that humans expect, use list.sort(key=sort_human) or sorted(list, key=sort_human)).

    :param _nsre:
    :param s: a python list
    :type s: list()
    :rtype: list()
    :return: liste human sort

    Example:
        >>> listToSorted = ["something1","something32","something17","something2","something29","something24"]
        >>> print(listToSorted.sort(key=sort_human))
        ['something1', 'something17', 'something2', 'something25', 'something29', 'something32']
        >>> print(sorted(listToSorted, key=sort_human))
        ['something1', 'something17', 'something2', 'something25', 'something29', 'something32']

    """
    try:
        return [int(text) if text.isdigit() else text.lower() for text in re.split(_nsre, s)]
    except TypeError:
        if not isinstance(s, int):
            print("WARNNING MODULES_SEB::sort_human : List %s value not understand so don't sort \n" % s)
        return s


def compareList(list1, list2):
    """
    Function to compare two list and return common, uniq1 and uniq2

    :param list1: a python list
    :type list1: list()
    :param list2: a python list
    :type list2: list()
    :return: list(), list(), list()
    :rtype: common, u1, u2
    :note: ens1 = set([1, 2, 3, 4, 5, 6])
    :note: ens2 = set([2, 3, 4])
    :note: ens3 = set([6, 7, 8, 9])
    :note: print ens1 & ens2 set([2, 3, 4]) car ce sont les seuls à être en même temps dans ens1 et ens2
    :note: print ens1 | ens3 set([1, 2, 3, 4, 5, 6, 7, 8, 9]), les deux réunis
    :note: print ens1 & ens3 set([6]), même raison que deux lignes au dessus
    :note: print ens1 ^ ens3 set([1, 2, 3, 4, 5, 7, 8, 9]), l'union moins les éléments communs
    :note: print ens1 - ens2 set([1, 5, 6]), on enlève les éléments de ens2


    Example:
        >>> l1 = [1, 2, 3, 4, 5, 6]
        >>> l2 = [6, 7, 8, 9]
        >>> com, u1, u2 = compareList(l1, l2)
        >>> print(com)
        [6]
        >>> print(u1)
        [1, 2, 3, 4, 5]
        >>> print(u2)
        [7, 8, 9]

    """

    ens1 = set(list1)
    ens2 = set(list2)
    common = list(ens1 & ens2)
    uniq1 = list(ens1 - ens2)
    uniq2 = list(ens2 - ens1)
    return sorted(common, key=sort_human), sorted(uniq1, key=sort_human), sorted(uniq2, key=sort_human)


def existant_file(x):
    """
    'Type' for argparse - checks that file exists but does not open by default.

    :param x: a file path
    :type x: str()
    :rtype: PosixPath
    :return: PosixPath

    """
    if not Path(x).exists():
        # Argparse uses the ArgumentTypeError to give a rejection message like:
        # error: argument input: x does not exist
        raise argparse.ArgumentTypeError(f'\n\nERROR: File "{x}" does not exist, please check file path')
    elif not Path(x).is_file():
        raise argparse.ArgumentTypeError(f'\n\nERROR: File "{x}" is not a valid file')

    return Path(x).resolve()


@Gooey(
        advanced=True,              # toggle whether to show advanced config or not
        show_config=True,           # skip config screens all together
        header_height=250,
        header_bg_color="#f0f0f0",
        default_size=(1000,700),    # starting size of the GUI
        fullscreen=False,
        dump_build_config=False,     # Dump the JSON Gooey uses to configure itself
        image_dir="./includes/",
        disable_stop_button=True,
        show_failure_modal=False,
        show_success_modal=True,
        richtext_controls=False,
        menu=[
                {
                'name' : 'Example',
                'items': [
                        {
                        'type'     : 'Link',
                        'menuTitle': 'Download example file',
                        'url'      : 'https://github.com/sravel/MLG_assign/raw/master/Test_MLG.xlsx'
                    }],
                },
                {
                'name' : 'Help',
                'items': [{
                        'type'     : 'Link',
                        'menuTitle': 'Documentation',
                        'url'      : 'https://github.com/sravel/MLG_assign'
                        },
                        {
                        'type'       : 'AboutDialog',
                        'menuTitle'  : 'About',
                        'name'       : 'MLG assign',
                        'description': 'Multi-Locus Genotypes assign',
                        'version'    : __version__,
                        'copyright'  : '2020',
                        'website'    : 'https://github.com/sravel/MLG_assign',
                        'developer'  : 'Sébastien RAVEL/',
                        'license'    : LICENSE
                        }
                    ]
                }
        ]
)
def GUI():
    parser = GooeyParser(prog='MLG_assign',
                         description=descriptionTools,
                         epilog=epilogTools
                         )

    inOutMandatory = parser.add_argument_group('Input mandatory infos for running', gooey_options={
            'columns': 1
    })

    inOutMandatory.add_argument('-x', '--excel', metavar="Excel filename", widget='FileChooser', type=existant_file,
                                required=True, dest='excel_file', help='matrice excel file')
    inOutMandatory.add_argument('-s', '--sheet', metavar="Name of Sheet", required=True, dest='sheet_name',
                                help='name of sheet in excel file')
    # inOutMandatory.add_argument('-d', '--debug', action='store_true', help='enter verbose/debug mode')

    args = parser.parse_args()

    return args

def main():

    parserOther = argparse.ArgumentParser(add_help=False)

    inOutOptional = parserOther.add_argument_group('Input infos not mandatory')
    inOutOptional.add_argument('-v', '--version', action='version', version=__version__,
                               help=f'Use if you want to know which version of {Path(__file__).stem} you are using')
    inOutOptional.add_argument('-h', '--help', action='help', help=f'show this help message and exit')
    # inOutOptional.add_argument('-d', '--debug', action='store_true', help='enter verbose/debug mode')

    parserMandatory = argparse.ArgumentParser(
            parents=[parserOther],
            add_help=False,
            prog=Path(__file__).stem,
            formatter_class=argparse.RawDescriptionHelpFormatter,
            description=descriptionTools,
            epilog=epilogTools
    )
    # Creer un sous parser
    subparsers = parserMandatory.add_subparsers(dest='mode', help=f'Select mode to run {Path(__file__).stem}')
    parserMandatory.set_defaults(mode="gui")
    # create the parser for the "cmd" command
    parser_cmd = subparsers.add_parser('cmd', help='active CMD mode')
    inOutMandatory = parser_cmd.add_argument_group('Input mandatory infos for running')
    inOutMandatory.add_argument('-e', '--excel', metavar="<path/to/file/Excel>", type=existant_file, required=True,
                                dest='excel_file', help='Matrix excel file')
    inOutMandatory.add_argument('-s', '--sheet', metavar="sheet name>", type=str, required=True,
                                dest='sheet_name', help='Name of sheet in excel file')
    # create the parser for the "gui" command
    parser_GUI = subparsers.add_parser('gui', help='enter GUI mode')
    parser_GUI.set_defaults()


    # Check parameters
    args = parserMandatory.parse_args()

    if args.mode == "gui":
        args = GUI()

    # Welcome message
    print(
            f"""{"#" * 80}\n#{Path(__file__).stem + " " + __version__:^78}#\n{"#" * 80}\nStart time: {datetime.now():%d-%m-%Y at %H:%M:%S}\nCommande line run: {" ".join(sys.argv)}\n""")
    # resume to user
    print(" - Intput info:")
    for k, v in vars(args).items():
        print(f"\t - {k}: {v}")


    # Récupère les arguments
    excel_file = args.excel_file
    sheet_name = args.sheet_name
    final_sheet_name = f"{args.sheet_name}_MLG"
    # debug = args.debug
    print(" - Output info:")
    print(f"\t - final sheet with MLG will be:{final_sheet_name}\n")
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:

            workBookbAll = load_workbook(excel_file)
            if sheet_name not in workBookbAll:
                raise NameError(f"\n\nERROR: Sheet {sheet_name} does not exist on file {excel_file}")
            if final_sheet_name in workBookbAll:
                raise NameError(f"\n\nERROR: MLG already exist, maybe script already run see sheet {final_sheet_name}")
            df = pd.read_excel(writer, sheet_name=sheet_name, index_col=0)
            df.MLG = ""

            # workBookbAll.remove(workBookbAll[sheet_name])
            # writer.book = workBookbAll

            dicoMLGUnique = {}  # clé n marqueurs valeur MLG
            dicoMissing = {}

            MLG = 1
            for row in df.itertuples(index=True, name='Pandas'):

                tabMicro = np.asarray(row[1:], dtype=np.int).tolist()
                ID = row[0]

                if 999 not in tabMicro:
                    tabMicroStr = ",".join([str(elm) for elm in tabMicro])

                    if tabMicroStr not in dicoMLGUnique.keys():
                        dicoMLGUnique[tabMicroStr] = MLG
                        MLG += 1
                    df.at[ID, "MLG"] = dicoMLGUnique[tabMicroStr]
                if 999 in list(tabMicro):
                    dicoMissing[ID] = tabMicro

            selectedMLG = {}
            dicomissingNewMLG = {}
            for ID, tabMicro in dicoMissing.items():
                values = np.array(tabMicro)
                indice = np.where(values == 999)[0]
                if ID not in selectedMLG.keys():
                    selectedMLG[ID] = []

                for tabMicroMLGUnique, MLGID in dicoMLGUnique.items():
                    newmicro = np.delete(values, indice)
                    newmicroMLG = np.delete(np.asarray(tabMicroMLGUnique.split(","), dtype=np.int), indice)

                    com, u1, u2 = compareList(newmicro.tolist(), newmicroMLG.tolist())
                    if len(u1) == 0 and len(u2) == 0:
                        selectedMLG[ID].append(MLGID)

                if len(selectedMLG[ID]) == 1:
                    df.at[ID, "MLG"] = dicoMLGUnique[tabMicroMLGUnique]

                if len(selectedMLG[ID]) == 0:

                    if str(newmicro) in dicomissingNewMLG.keys():
                        df.at[ID, "MLG"] = dicomissingNewMLG[str(newmicro)]
                    else:
                        dicomissingNewMLG[str(newmicro)] = MLG
                        df.at[ID, "MLG"] = MLG
                        MLG += 1
                else:
                    df.at[ID, "MLG"] = "NaN"
            # if debug:
            print(f"Final matrix header:\n{df.head(30)}")
            df.to_excel(writer, sheet_name=final_sheet_name, merge_cells=False, na_rep="nd")
            writer.save()
    except NameError as e:
        print(e)
        sys.exit(1)
    except KeyError as e:
        print(f"\n\nERROR: please check if file is not open: {excel_file}\n{e}")
        sys.exit(1)

    print(f"""\nStop time: {datetime.now():%d-%m-%Y at %H:%M:%S}\n{'#' * 80}\n#{'End of execution':^78}#\n{"#" * 80}""")


###################################################
# Main code
##################################################
if __name__ == "__main__":
    main()